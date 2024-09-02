import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
filename = "C:\\Users\\HP\\OneDrive\\Desktop\\python\\IEEE CS website\\flights-dataset.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]
n_arr=[]
print(wb.sheetnames)
worksheet1 = wb['Sheet1']
flight_no = []
for row in worksheet1.iter_rows():
    flight_no.append(row[2].value)
for i in flight_no[2:]:
    p=str(i).strip('\n')
    n_arr.append(p)
i = 0
j = 0
for aircraft in n_arr:
    if "\n" in aircraft:
        j += 1
        pass
    else:
        html_text = requests.get(f"https://www.delhiairport.com/igi-indira-gandhi-flight-arrival/{aircraft}").text
        soup = BeautifulSoup(html_text, 'lxml')
        aircraft_type = soup.find_all("div", class_="flight-airline__text")
        try:
            ws["F"+str(i)] = aircraft_type[1].text
            wb.save(filename)
           #print(aircraft_type[1].text)
        except IndexError:
            ws["F"+str(i)] = "NA"
            wb.save(filename)
        j += 1
    print("excel index {}".format(i))
    print(f"aircraft index {j}")
       
    i += 1
